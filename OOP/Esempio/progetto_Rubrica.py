import tkinter as tk
from openpyxl import Workbook, load_workbook
import os
from tkinter import messagebox

# Nome del file excel
FILE_EXCEL = "rubrica_contatti.xlsx"

# Verifica se il file esiste, crea con intestazione
if not os.path.exists(FILE_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatti"
    ws.append(["Nome", "Cognome", "Data di Nascita", "Numero di Telefono"])
    wb.save(FILE_EXCEL)

root = tk.Tk()

# Titolo del form
root.title("Rubrica Contatti")

# Dimensione e sfondo
root.geometry("400x400")
root.configure(bg="#0A8D09")

# Creazione Label + entry per Nome
tk.Label(root, text = 'Nome: ', font = ("Roboto", 15), bg="#0A8D09", fg="#ffffff", anchor="w").pack(fill="x", padx=20)
entry_nome = tk.Entry(root, width=30, font=("Roboto", 15))
entry_nome.pack(pady=5)

# Creazione Label + entry per Cognome
tk.Label(root, text = 'Cognome: ', font = ("Roboto", 15), bg="#0A8D09", fg="#ffffff", anchor="w").pack(fill="x", padx=20)
entry_cognome = tk.Entry(root, width=30, font=("Roboto", 15))
entry_cognome.pack(pady=5)

# Creazione Label + entry per Data Nascita
tk.Label(root, text = 'Data Nascita (gg/mm/yyyy): ', font = ("Roboto", 15), bg="#0A8D09", fg="#ffffff", anchor="w").pack(fill="x", padx=20)
entry_data_nascita = tk.Entry(root, width=30, font=("Roboto", 15))
entry_data_nascita.pack(pady=5)

# Creazione Label + entry per Numero di Telefono
tk.Label(root, text = 'Numero di Telefono: ', font = ("Roboto", 15), bg="#0A8D09", fg="#ffffff", anchor="w").pack(fill="x", padx=20)
entry_telefono = tk.Entry(root, width=30, font=("Roboto", 15))
entry_telefono.pack(pady=5)

# FUNZIONE PER SALVARE I DATI
def salva_dati():
    nome = entry_nome.get()
    cognome = entry_cognome.get()
    data = entry_data_nascita.get()
    telefono = entry_telefono.get()

    wb = load_workbook(FILE_EXCEL)
    ws = wb.active
    ws.append([nome, cognome, data, telefono])
    wb.save(FILE_EXCEL)

    messagebox.showinfo(
        "Info Salvataggio", 
        "Il salvataggio è andato a buon fine :D"
    )

    
    svuotaCampi()

def svuotaCampi():
    entry_nome.delete(0, tk.END)
    entry_cognome.delete(0, tk.END)
    entry_data_nascita.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)

# BOTTONE SALVA
button = tk.Button(
    root,
    text='Salva Contatto',
    font=("Roboto", 15, "bold"),
    bg="#ffffff",
    command=salva_dati
)
button.pack(pady=20)


root.mainloop()